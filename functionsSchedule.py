from itertools import chain

import openpyxl
from openpyxl import load_workbook
import datetime
import re

def date_check(date_str_checking):
    if '..' in date_str_checking:
        date_str_checking=date_str_checking.replace('..','.')

    date_line = re.search('\d{2}-\d{2}-\d{4}|\d{2}.\d{2}.\d{4}|\d{2}.\d{2}.\d{2}|\d{2}..\d{2}.\d{4}',
                              date_str_checking)
    date = datetime.datetime.strptime(date_line.group(), '%d.%m.%Y').date()
    return date


def processing(filename):
    dict_teach_and_group = {}
    week = ""
    day = ""
    val = None
    weeks = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
    workbook = load_workbook(filename)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        group_row = None
        for row in sheet.rows:
            if row[0].value is not None and row[0].value.split()[0] in weeks:
                week = row[0].value
                (week, day) = week.split()
            elif not  isinstance(row[0], openpyxl.cell.cell.MergedCell):
                if row[1].value == "Время":
                    # if not any(val.value is None for val in row):
                    group_row = row
                continue

            time = row[1].value
            if datetime.date.today() <= date_check(day):
                week_day_time = (week, day, time)
                dict_teach_and_group[week_day_time] = {}
                for i in range(2, len(row)):
                    group = group_row[i].value
                    group = str(group)
                    if 'Общая физическая подготовка пз'  in str(row[i].value):
                        break
                    if isinstance(row[i], openpyxl.cell.cell.MergedCell) and val is not None:
                        if val not in dict_teach_and_group[week_day_time]:
                            dict_teach_and_group[week_day_time][val] = [group.strip()]
                        else:
                            dict_teach_and_group[week_day_time][val] += [group.strip()]
                    elif row[i].value is not None:
                        val = row[i].value
                        val = ' '.join(val.split())
                        if val not in dict_teach_and_group[week, day, time]:
                            dict_teach_and_group[week_day_time][val] = [group.strip()]
                        else:
                            dict_teach_and_group[week_day_time][val] += [group.strip()]
                    else:
                        val = None
    return dict_teach_and_group





def toStr(timedate, subgroups,form, typeT, flag=False):
    timedate = " ".join(timedate)
    name = list(subgroups.keys())[0]
    begin = '\n' if flag else ''
    stringForAdding = begin + '<b>' + str(timedate) + '\n' + "предмет: " + str(name) \
                      + '\n' + f"{form}({typeT}) группы: " \
                      + str(', '.join(subgroups[name])) + '\n\n</b>'
    return stringForAdding

def checkForm(file):
    return 'Магистратура' if 'mag' in file else 'Бакалавриат'


def printing_schedule(flag):
    if flag == 'очное':
        flist = ['1kyrs.xlsx', '2kyrs.xlsx', '3kyrs.xlsx', '4kyrs.xlsx', 'mag1.xlsx', 'mag2.xlsx']
    elif flag == 'очно-заочное':
        flist = ['1kyrsOZ.xlsx','2kyrsOZ.xlsx','4kyrsOZ.xlsx','mag1oz.xlsx']
    else:
        flist = []
    lists = [[y+(checkForm(x),) for y in processing(x).items()] for x in flist]
    vaiues = list(chain(*lists))
    vaiues = [(x[0], {z: x[1][z] for z in x[1] if 'Янгирова' in z}, x[2]) for x in vaiues if
              any([True for y in x[1] if 'Янгирова' in y])]

    vaiues = [x for x in vaiues if
              datetime.date.today() <= date_check(x[0][1]) <= datetime.date.today() + datetime.timedelta(days=31)]
    vaiues = sorted(vaiues, key=lambda x: date_check(x[0][1]))
    day = None
    newVaiues = ""
    for vaiue in vaiues:
        newVaiues += toStr(*vaiue, flag, day != vaiue[0][1])
        day = vaiue[0][1]
    return newVaiues
