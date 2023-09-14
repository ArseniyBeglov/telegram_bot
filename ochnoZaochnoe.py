import openpyxl
from openpyxl import load_workbook
import datetime
import re




def processing(filename):
    dict_teach_and_group = {}
    week = ""
    val = None

    workbook = load_workbook(filename)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        group_row = None
        for row in sheet.rows:
            if row[1].value == "Время":
                # if not any(val.value is None for val in row):
                group_row = row
                continue
            if group_row is None:
                continue
            if row[0].value is not None:
                week = row[0].value
                week = ' '.join(week.split())
            if row[1].value is None:
                continue
            time = row[1].value

            key_time = week + " " + time
            dict_teach_and_group[key_time] = {}

            for i in range(2, len(row)):
                group = group_row[i].value
                group = str(group)
                if isinstance(row[i], openpyxl.cell.cell.MergedCell) and val is not None:
                    dict_teach_and_group[key_time][val] += [group.strip()]
                elif row[i].value is not None:
                    val = row[i].value
                    val = ' '.join(val.split())
                    if val not in dict_teach_and_group[key_time]:
                        dict_teach_and_group[key_time][val] = [group.strip()]
                    else:
                        dict_teach_and_group[key_time][val] += [group.strip()]
                else:
                    val = None
    return dict_teach_and_group


def date_check(date_str_checking):
    date_line = re.search('\d{2}-\d{2}-\d{4}|\d{2}.\d{2}.\d{4}|\d{2}.\d{2}.\d{2}|\d{2}..\d{2}.\d{4}', date_str_checking)
    date = datetime.datetime.strptime(date_line.group(), '%d.%m.%Y').date()
    return date



def cheking_time(key, list_day):
    if '08:00' in key:
        list_day.insert(0, key)
    elif '09:35' in key:
        list_day.insert(1, key)
    elif '11:35' in key:
        list_day.insert(2, key)
    elif '13:10' in key:
        list_day.insert(3, key)
    elif '15:10' in key:
        list_day.insert(4, key)
    elif '16:45' in key:
        list_day.insert(5, key)
    elif '18:20' in key:
        list_day.insert(6, key)
    elif '19:55' in key:
        list_day.insert(7, key)
    return list_day


def printing_schedule():
    string_monday = ['', '', '', '', '', '', '', '']
    string_tuesday = ['', '', '', '', '', '', '', '']
    string_wednesday = ['', '', '', '', '', '', '', '']
    string_thursday = ['', '', '', '', '', '', '', '']
    string_friday = ['', '', '', '', '', '', '', '']
    string_saturday = ['', '', '', '', '', '', '', '']
    flist = ['4kyrsOZ.xlsx']
    for fname in flist:
        mergeddicts = processing(fname)
        dd=None
        for key in mergeddicts:
            for name in mergeddicts[key]:
                if 'Янгирова' in name:
                    if datetime.date.today() <= date_check(key) :
                        if dd ==date_check(key):
                            string_for_adding = str(key) + '\n' + "предмет: " + str(name) \
                                                + '\n' + "Бакалавриат(очно-заочное) группы: " \
                                                + str(', '.join(mergeddicts[key][name])) + " " + '\n'+ '\n'
                        else: string_for_adding = str(key) + '\n' + "предмет: " + str(name) \
                                                + '\n' + "Бакалавриат(очно-заочное) группы: " \
                                                + str(', '.join(mergeddicts[key][name])) + " " + '\n'
                        dd=date_check(key)
                        if 'Понедельник' in key:
                            string_monday.append(string_for_adding)
                        elif 'Вторник' in key:
                            string_tuesday.append(string_for_adding)
                        elif 'Среда' in key:
                            string_wednesday.append(string_for_adding)
                        elif 'Четверг' in key:
                            string_thursday.append(string_for_adding)
                        elif 'Пятница' in key:
                            string_friday.append(string_for_adding)
                        elif 'Суббота' in key:
                            string_saturday.append(string_for_adding)
    string_schedule = str(''.join(string_monday)) + '\n' + '\n' + str(''.join(string_tuesday)) + '\n' + '\n' + str(
        ''.join(string_wednesday)) \
                      + '\n' + '\n' + \
                      str(''.join(string_thursday)) + '\n' + '\n' + str(''.join(string_friday)) + '\n' + '\n' + str(
        ''.join(string_saturday))
    return string_schedule
